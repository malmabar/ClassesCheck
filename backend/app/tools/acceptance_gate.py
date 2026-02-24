from __future__ import annotations

import argparse
import csv
import json
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib import error as url_error
from urllib import parse as url_parse
from urllib import request as url_request

import openpyxl

from app.core.config import PROJECT_ROOT


PERIODS: Tuple[str, str] = ("صباحي", "مسائي")
DAY_ORDER_MAP: Dict[str, int] = {
    "الاحد": 1,
    "الأحد": 1,
    "الاحد ": 1,
    "الاثنين": 2,
    "الإثنين": 2,
    "الثلاثاء": 3,
    "الاربعاء": 4,
    "الأربعاء": 4,
    "الخميس": 5,
}
SCHEDULE_BY_PERIOD: Dict[str, Tuple[Tuple[int, int], ...]] = {
    "صباحي": (
        (800, 850),
        (900, 950),
        (1000, 1050),
        (1100, 1140),
        (1230, 1320),
        (1321, 1410),
        (1415, 1505),
        (1506, 1556),
    ),
    "مسائي": (
        (1600, 1650),
        (1651, 1741),
        (1750, 1840),
        (1841, 1931),
        (1940, 2030),
        (2031, 2121),
        (2130, 2220),
        (2221, 2311),
    ),
}


@dataclass
class GateOptions:
    base_url: str
    workbook_file: Path
    source_csv: Optional[Path]
    semester: Optional[str]
    period: str
    created_by: str
    timeout_sec: int
    output_file: Path
    save_timestamped: bool


def parse_args() -> GateOptions:
    parser = argparse.ArgumentParser(
        description="Acceptance regression gate for Morning Classes Check (distribution + publish + exports).",
    )
    parser.add_argument(
        "--base-url",
        default="http://127.0.0.1:8000",
        help="API base URL (default: http://127.0.0.1:8000).",
    )
    parser.add_argument(
        "--workbook-file",
        default=str(PROJECT_ROOT / "MorningClassesCheck - Beta6.xlsm"),
        help="Workbook path used for source extraction and evening Excel parity.",
    )
    parser.add_argument(
        "--source-csv",
        default="",
        help="Optional SS01 CSV file. If omitted, script extracts SS01_Report from workbook.",
    )
    parser.add_argument(
        "--semester",
        default="",
        help="Optional semester override. If omitted, derived from workbook SS01_Report!A2.",
    )
    parser.add_argument(
        "--period",
        choices=("all", "صباحي", "مسائي"),
        default="all",
        help="Run gate for one period or both (default: all).",
    )
    parser.add_argument(
        "--created-by",
        default="acceptance-gate-bot",
        help="Operator value sent to API endpoints.",
    )
    parser.add_argument(
        "--timeout-sec",
        type=int,
        default=120,
        help="HTTP timeout in seconds (default: 120).",
    )
    parser.add_argument(
        "--output-file",
        default=str(PROJECT_ROOT / "artifacts" / "acceptance" / "latest.json"),
        help="Output report path (default: artifacts/acceptance/latest.json).",
    )
    parser.add_argument(
        "--save-timestamped",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Also write artifacts/acceptance/acceptance_<timestamp>.json (default: true).",
    )
    args = parser.parse_args()

    workbook_file = Path(args.workbook_file).expanduser().resolve()
    if not workbook_file.exists():
        parser.error(f"Workbook file does not exist: {workbook_file}")

    source_csv = Path(args.source_csv).expanduser().resolve() if args.source_csv else None
    if source_csv and not source_csv.exists():
        parser.error(f"source CSV does not exist: {source_csv}")

    output_file = Path(args.output_file).expanduser().resolve()
    output_file.parent.mkdir(parents=True, exist_ok=True)

    semester = str(args.semester).strip() or None
    return GateOptions(
        base_url=args.base_url.rstrip("/"),
        workbook_file=workbook_file,
        source_csv=source_csv,
        semester=semester,
        period=args.period,
        created_by=args.created_by,
        timeout_sec=max(10, int(args.timeout_sec)),
        output_file=output_file,
        save_timestamped=bool(args.save_timestamped),
    )


def _decode_csv(file_bytes: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "cp1256"):
        try:
            return file_bytes.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise ValueError("Unable to decode CSV. Supported encodings: UTF-8, Windows-1256.")


def _detect_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:20])
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;").delimiter
    except csv.Error:
        return ","


def _extract_ss01_report_to_csv(workbook_file: Path, output_csv: Path) -> dict:
    wb = openpyxl.load_workbook(workbook_file, data_only=True, keep_vba=True)
    ws = wb["SS01_Report"]
    headers = [ws.cell(1, c).value for c in range(1, 32)]
    if not headers or str(headers[0] or "").strip() != "الفصل التدريبي":
        raise ValueError("SS01_Report header row is invalid in workbook.")

    rows_written = 0
    with output_csv.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for row_idx in range(2, ws.max_row + 1):
            values = [ws.cell(row_idx, c).value for c in range(1, 32)]
            if all(v in (None, "") for v in values):
                continue
            writer.writerow(["" if v is None else v for v in values])
            rows_written += 1
    semester = str(ws["A2"].value).strip() if ws["A2"].value not in (None, "") else None
    return {"rows_written": rows_written, "semester": semester}


def _normalize_digit_chars(value: Any) -> str:
    mapping = {
        "٠": "0",
        "١": "1",
        "٢": "2",
        "٣": "3",
        "٤": "4",
        "٥": "5",
        "٦": "6",
        "٧": "7",
        "٨": "8",
        "٩": "9",
        "۰": "0",
        "۱": "1",
        "۲": "2",
        "۳": "3",
        "۴": "4",
        "۵": "5",
        "۶": "6",
        "۷": "7",
        "۸": "8",
        "۹": "9",
    }
    return "".join(mapping.get(ch, ch) for ch in str(value or ""))


def _hhmm_to_minutes(hhmm: Optional[int]) -> Optional[int]:
    if not isinstance(hhmm, int):
        return None
    hours = hhmm // 100
    minutes = hhmm % 100
    if hours < 0 or hours > 23 or minutes < 0 or minutes > 59:
        return None
    return hours * 60 + minutes


def _parse_token_to_hhmm(token: str) -> Optional[int]:
    text = _normalize_digit_chars(token).strip()
    if not text:
        return None
    if ":" in text or "." in text:
        sep = ":" if ":" in text else "."
        left, right = text.split(sep, 1)
        if left.isdigit() and right.isdigit():
            hours = int(left)
            minutes = int(right)
            if 0 <= hours <= 23 and 0 <= minutes <= 59:
                return (hours * 100) + minutes
            return None
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) == 3:
        hours = int(digits[0])
        minutes = int(digits[1:])
        if 0 <= hours <= 23 and 0 <= minutes <= 59:
            return (hours * 100) + minutes
    if len(digits) == 4:
        hours = int(digits[:2])
        minutes = int(digits[2:])
        if 0 <= hours <= 23 and 0 <= minutes <= 59:
            return (hours * 100) + minutes
    return None


def _extract_hhmm_tokens(text_value: Any) -> List[int]:
    normalized = _normalize_digit_chars(text_value)
    tokens: List[int] = []
    current = ""
    for ch in normalized:
        if ch.isdigit() or ch in (":", "."):
            current += ch
        else:
            if current:
                parsed = _parse_token_to_hhmm(current)
                if isinstance(parsed, int):
                    tokens.append(parsed)
                current = ""
    if current:
        parsed = _parse_token_to_hhmm(current)
        if isinstance(parsed, int):
            tokens.append(parsed)
    return tokens


def _resolve_time_range(time_value: Any, time_hhmm: Optional[int] = None) -> dict:
    tokens = _extract_hhmm_tokens(time_value)
    first = tokens[0] if tokens else None
    if not isinstance(first, int) and isinstance(time_hhmm, int):
        first = time_hhmm
    second = tokens[1] if len(tokens) > 1 else None
    start_hhmm = first if isinstance(first, int) else None
    end_hhmm = second if isinstance(second, int) else None
    start_min = _hhmm_to_minutes(start_hhmm)
    end_min = _hhmm_to_minutes(end_hhmm)
    if start_min is not None and end_min is not None and end_min < start_min:
        start_hhmm, end_hhmm = end_hhmm, start_hhmm
        start_min, end_min = end_min, start_min
    return {
        "start_hhmm": start_hhmm,
        "end_hhmm": end_hhmm,
        "start_min": start_min,
        "end_min": end_min,
    }


def _collect_overlapping_slots(period: str, time_range: dict) -> List[int]:
    start_min = time_range["start_min"]
    if start_min is None:
        return []
    end_min = time_range["end_min"]
    effective_end = end_min if end_min is not None and end_min > start_min else start_min + 1
    slots: List[int] = []
    for index, (start_hhmm, end_hhmm) in enumerate(SCHEDULE_BY_PERIOD[period], start=1):
        slot_start = _hhmm_to_minutes(start_hhmm)
        slot_end = _hhmm_to_minutes(end_hhmm)
        if slot_start is None or slot_end is None:
            continue
        if start_min < slot_end and effective_end > slot_start:
            slots.append(index)
    return sorted(set(slots))


def _find_slot_by_time(period: str, hhmm: Optional[int]) -> Optional[int]:
    if not isinstance(hhmm, int):
        return None
    for idx, (start_hhmm, _end_hhmm) in enumerate(SCHEDULE_BY_PERIOD[period], start=1):
        if hhmm == start_hhmm:
            return idx
    minutes = _hhmm_to_minutes(hhmm)
    if minutes is None:
        return None
    for idx, (start_hhmm, end_hhmm) in enumerate(SCHEDULE_BY_PERIOD[period], start=1):
        start = _hhmm_to_minutes(start_hhmm)
        end = _hhmm_to_minutes(end_hhmm)
        if start is None or end is None:
            continue
        if minutes >= start and minutes < end:
            return idx
    return None


def _infer_period(time_range: dict, section_type: str, period_hint: Optional[str]) -> str:
    morning_slots = _collect_overlapping_slots("صباحي", time_range)
    evening_slots = _collect_overlapping_slots("مسائي", time_range)
    if len(morning_slots) != len(evening_slots):
        return "صباحي" if len(morning_slots) > len(evening_slots) else "مسائي"
    start_hhmm = time_range["start_hhmm"]
    end_hhmm = time_range["end_hhmm"]
    reference = start_hhmm if isinstance(start_hhmm, int) else end_hhmm
    if isinstance(reference, int):
        return "مسائي" if reference >= 1600 else "صباحي"
    section_text = str(section_type or "")
    if "صباح" in section_text:
        return "صباحي"
    if "مسائ" in section_text:
        return "مسائي"
    if period_hint in PERIODS:
        return period_hint
    return "صباحي"


def _resolve_slot_indices(row: dict, period: str, time_range: dict) -> List[int]:
    slots = _collect_overlapping_slots(period, time_range)
    alt_period = "مسائي" if period == "صباحي" else "صباحي"
    alt_slots = _collect_overlapping_slots(alt_period, time_range)
    if len(alt_slots) > len(slots):
        slots = alt_slots

    if not slots and isinstance(time_range["start_hhmm"], int):
        by_period = _find_slot_by_time(period, time_range["start_hhmm"])
        by_alt = _find_slot_by_time(alt_period, time_range["start_hhmm"])
        if isinstance(by_period, int):
            slots = [by_period]
        elif isinstance(by_alt, int):
            slots = [by_alt]

    if not slots:
        fallback_slot = row.get("slot_index")
        if isinstance(fallback_slot, int) and 1 <= fallback_slot <= 8:
            slots = [fallback_slot]

    return sorted(set(slots))


def _normalize_day_order(day_order: Any, day_name: Any) -> Optional[int]:
    if isinstance(day_order, int) and 1 <= day_order <= 5:
        return day_order
    normalized = str(day_name or "").strip()
    if not normalized:
        return None
    return DAY_ORDER_MAP.get(normalized) or DAY_ORDER_MAP.get(normalized.replace(" ", ""))


def _trainer_key_from_row(row: dict, id_key: str, name_key: str) -> str:
    trainer_id = str(row.get(id_key) or "").strip()
    if trainer_id:
        return trainer_id
    trainer_name = str(row.get(name_key) or "").strip()
    if trainer_name:
        return trainer_name
    return ""


def _iter_csv_rows(csv_file: Path) -> tuple[Iterable[dict], dict]:
    payload = csv_file.read_bytes()
    decoded_text, encoding = _decode_csv(payload)
    delimiter = _detect_delimiter(decoded_text)
    reader = csv.DictReader(decoded_text.splitlines(), delimiter=delimiter)
    rows = list(reader)
    return rows, {"encoding": encoding, "delimiter": delimiter, "row_count": len(rows)}


def _build_expected_cell_counts(csv_file: Path, period: str) -> dict:
    rows, meta = _iter_csv_rows(csv_file)
    cell_trainers: Dict[Tuple[int, int], set[str]] = {(d, s): set() for d in range(1, 6) for s in range(1, 9)}
    source_rows_used = 0

    for row in rows:
        time_range = _resolve_time_range(row.get("الوقت"), None)
        resolved_period = _infer_period(time_range, str(row.get("نوع الشعبة") or ""), period)
        if resolved_period != period:
            continue
        day_order = _normalize_day_order(None, row.get("اليوم"))
        if day_order is None:
            continue
        trainer_key = _trainer_key_from_row(row, "رقم المدرب", "اسم المدرب")
        if not trainer_key:
            continue
        slots = _resolve_slot_indices({}, resolved_period, time_range)
        if not slots:
            continue
        source_rows_used += 1
        for slot_idx in slots:
            if 1 <= slot_idx <= 8:
                cell_trainers[(day_order, slot_idx)].add(trainer_key)

    counts = {(d, s): len(cell_trainers[(d, s)]) for d in range(1, 6) for s in range(1, 9)}
    return {
        "counts": counts,
        "source_rows_used": source_rows_used,
        "total_loads": int(sum(counts.values())),
        "csv_meta": meta,
    }


def _build_actual_cell_counts(base_url: str, run_id: str, period: str, timeout_sec: int) -> dict:
    page = 1
    size = 500
    items: List[dict] = []
    while True:
        payload = _http_json(
            method="GET",
            url=f"{base_url}/api/v1/mc/runs/{run_id}/codes?page={page}&size={size}",
            timeout_sec=timeout_sec,
        )
        rows = payload.get("items") or []
        items.extend(rows)
        if not payload.get("has_next"):
            break
        page += 1

    cell_trainers: Dict[Tuple[int, int], set[str]] = {(d, s): set() for d in range(1, 6) for s in range(1, 9)}
    rows_used = 0
    for row in items:
        day_order = _normalize_day_order(row.get("day_order"), row.get("day_name"))
        if day_order is None:
            continue
        trainer_key = _trainer_key_from_row(row, "trainer_job_id", "trainer_name")
        if not trainer_key:
            continue
        time_range = _resolve_time_range(row.get("time_value"), row.get("time_hhmm"))
        section_type = str(row.get("section_type") or "")
        resolved_period = _infer_period(time_range, section_type, row.get("period"))
        if resolved_period != period:
            period_flag = bool(row.get("is_morning")) if period == "صباحي" else bool(row.get("is_evening"))
            if not period_flag:
                continue
        slots = _resolve_slot_indices(row, period, time_range)
        if not slots:
            continue
        rows_used += 1
        for slot_idx in slots:
            if 1 <= slot_idx <= 8:
                cell_trainers[(day_order, slot_idx)].add(trainer_key)

    counts = {(d, s): len(cell_trainers[(d, s)]) for d in range(1, 6) for s in range(1, 9)}
    return {
        "counts": counts,
        "codes_rows": len(items),
        "rows_used": rows_used,
        "total_loads": int(sum(counts.values())),
    }


def _compare_count_maps(expected: dict, actual: dict) -> dict:
    mismatches = []
    for day_order in range(1, 6):
        for slot_idx in range(1, 9):
            key = (day_order, slot_idx)
            exp_value = int(expected[key])
            act_value = int(actual[key])
            if exp_value != act_value:
                mismatches.append(
                    {
                        "day_order": day_order,
                        "slot_index": slot_idx,
                        "expected": exp_value,
                        "actual": act_value,
                        "delta": act_value - exp_value,
                    }
                )
    return {
        "mismatch_count": len(mismatches),
        "mismatches_sample": mismatches[:20],
    }


def _calculate_distribution_ratios_from_counts(counts: dict) -> dict:
    total = float(sum(counts.values()))
    weekly = {(d, s): ((counts[(d, s)] / total) if total > 0 else 0.0) for d in range(1, 6) for s in range(1, 9)}
    day_totals = {d: sum(counts[(d, s)] for s in range(1, 9)) for d in range(1, 6)}
    daily = {
        (d, s): ((counts[(d, s)] / day_totals[d]) if day_totals[d] > 0 else None)
        for d in range(1, 6)
        for s in range(1, 9)
    }
    slot_dist = {s: sum(weekly[(d, s)] for d in range(1, 6)) for s in range(1, 9)}
    return {"weekly": weekly, "daily": daily, "slot_dist": slot_dist, "day_totals": day_totals, "total": total}


def _compare_with_evening_excel_cache(workbook_file: Path, counts: dict, period: str, tol: float = 1e-9) -> dict:
    wb = openpyxl.load_workbook(workbook_file, data_only=True, keep_vba=True)
    ws = wb["التوزيع النسبي"]
    title = str(ws["B1"].value or "")
    if period not in title:
        return {
            "checked": False,
            "reason": f"Workbook cache title is '{title}', not '{period}'.",
            "mismatch_count": 0,
            "mismatches_sample": [],
        }

    ratios = _calculate_distribution_ratios_from_counts(counts)
    weekly = ratios["weekly"]
    daily = ratios["daily"]
    slot_dist = ratios["slot_dist"]
    day_totals = ratios["day_totals"]
    total = ratios["total"]

    start_col = {1: 2, 2: 10, 3: 18, 4: 26, 5: 34}
    row20_cols = {1: 2, 2: 7, 3: 12, 4: 17, 5: 22, 6: 27, 7: 32, 8: 37}
    mismatches = []

    def _float(value: Any) -> Optional[float]:
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    for day_order in range(1, 6):
        weekly_values = []
        daily_values = []
        for slot_idx in range(1, 9):
            col = start_col[day_order] + (slot_idx - 1)
            cell_count = int(ws.cell(5, col).value or 0)
            expected_count = int(counts[(day_order, slot_idx)])
            if cell_count != expected_count:
                mismatches.append(
                    {
                        "cell": ws.cell(5, col).coordinate,
                        "metric": "count",
                        "excel": cell_count,
                        "expected": expected_count,
                    }
                )
            cell_weekly = _float(ws.cell(6, col).value) or 0.0
            expected_weekly = weekly[(day_order, slot_idx)]
            if abs(cell_weekly - expected_weekly) > tol:
                mismatches.append(
                    {
                        "cell": ws.cell(6, col).coordinate,
                        "metric": "weekly_ratio",
                        "excel": cell_weekly,
                        "expected": expected_weekly,
                    }
                )
            cell_daily = ws.cell(14, col).value
            expected_daily = daily[(day_order, slot_idx)]
            if expected_daily is None:
                if cell_daily not in (None, "لا"):
                    mismatches.append(
                        {
                            "cell": ws.cell(14, col).coordinate,
                            "metric": "daily_ratio",
                            "excel": cell_daily,
                            "expected": "لا",
                        }
                    )
                daily_values.append(0.0)
            else:
                excel_daily = _float(cell_daily)
                if excel_daily is None or abs(excel_daily - expected_daily) > tol:
                    mismatches.append(
                        {
                            "cell": ws.cell(14, col).coordinate,
                            "metric": "daily_ratio",
                            "excel": cell_daily,
                            "expected": expected_daily,
                        }
                    )
                daily_values.append(expected_daily)
            weekly_values.append(expected_weekly)

        left_col = start_col[day_order]
        right_col = start_col[day_order] + 4
        excel_week_half_1 = _float(ws.cell(7, left_col).value) or 0.0
        excel_week_half_2 = _float(ws.cell(7, right_col).value) or 0.0
        exp_week_half_1 = sum(weekly_values[:4])
        exp_week_half_2 = sum(weekly_values[4:])
        if abs(excel_week_half_1 - exp_week_half_1) > tol:
            mismatches.append(
                {"cell": ws.cell(7, left_col).coordinate, "metric": "weekly_half_1", "excel": excel_week_half_1, "expected": exp_week_half_1}
            )
        if abs(excel_week_half_2 - exp_week_half_2) > tol:
            mismatches.append(
                {"cell": ws.cell(7, right_col).coordinate, "metric": "weekly_half_2", "excel": excel_week_half_2, "expected": exp_week_half_2}
            )

        excel_day_total = _float(ws.cell(8, left_col).value) or 0.0
        exp_day_total = sum(weekly_values)
        if abs(excel_day_total - exp_day_total) > tol:
            mismatches.append(
                {"cell": ws.cell(8, left_col).coordinate, "metric": "weekly_day_total", "excel": excel_day_total, "expected": exp_day_total}
            )

        excel_daily_half_1 = _float(ws.cell(15, left_col).value) or 0.0
        excel_daily_half_2 = _float(ws.cell(15, right_col).value) or 0.0
        exp_daily_half_1 = sum(daily_values[:4])
        exp_daily_half_2 = sum(daily_values[4:])
        if abs(excel_daily_half_1 - exp_daily_half_1) > tol:
            mismatches.append(
                {"cell": ws.cell(15, left_col).coordinate, "metric": "daily_half_1", "excel": excel_daily_half_1, "expected": exp_daily_half_1}
            )
        if abs(excel_daily_half_2 - exp_daily_half_2) > tol:
            mismatches.append(
                {"cell": ws.cell(15, right_col).coordinate, "metric": "daily_half_2", "excel": excel_daily_half_2, "expected": exp_daily_half_2}
            )

        excel_day_share = _float(ws.cell(16, left_col).value) or 0.0
        exp_day_share = (day_totals[day_order] / total) if total > 0 else 0.0
        if abs(excel_day_share - exp_day_share) > tol:
            mismatches.append(
                {"cell": ws.cell(16, left_col).coordinate, "metric": "day_share", "excel": excel_day_share, "expected": exp_day_share}
            )

    for slot_idx, col in row20_cols.items():
        excel_slot = _float(ws.cell(20, col).value) or 0.0
        exp_slot = slot_dist[slot_idx]
        if abs(excel_slot - exp_slot) > tol:
            mismatches.append(
                {"cell": ws.cell(20, col).coordinate, "metric": "slot_distribution", "excel": excel_slot, "expected": exp_slot}
            )

    excel_first = _float(ws.cell(21, 2).value) or 0.0
    excel_second = _float(ws.cell(21, 22).value) or 0.0
    exp_first = sum(slot_dist[idx] for idx in (1, 2, 3, 4))
    exp_second = sum(slot_dist[idx] for idx in (5, 6, 7, 8))
    if abs(excel_first - exp_first) > tol:
        mismatches.append({"cell": "B21", "metric": "slot_split_first", "excel": excel_first, "expected": exp_first})
    if abs(excel_second - exp_second) > tol:
        mismatches.append({"cell": "V21", "metric": "slot_split_second", "excel": excel_second, "expected": exp_second})

    return {
        "checked": True,
        "reason": "",
        "mismatch_count": len(mismatches),
        "mismatches_sample": mismatches[:20],
        "title_cell": title,
    }


def _http_json(method: str, url: str, timeout_sec: int, payload: Optional[bytes] = None, headers: Optional[dict] = None) -> dict:
    req = url_request.Request(url=url, method=method, data=payload, headers=headers or {})
    try:
        with url_request.urlopen(req, timeout=timeout_sec) as response:
            text = response.read().decode("utf-8", errors="replace")
            if not text.strip():
                return {}
            return json.loads(text)
    except url_error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code} {method} {url}: {detail}") from exc
    except url_error.URLError as exc:
        raise RuntimeError(f"Failed {method} {url}: {exc}") from exc


def _http_binary(method: str, url: str, timeout_sec: int) -> tuple[bytes, dict]:
    req = url_request.Request(url=url, method=method)
    try:
        with url_request.urlopen(req, timeout=timeout_sec) as response:
            data = response.read()
            return data, dict(response.headers.items())
    except url_error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code} {method} {url}: {detail}") from exc
    except url_error.URLError as exc:
        raise RuntimeError(f"Failed {method} {url}: {exc}") from exc


def _build_multipart(fields: dict[str, str], file_field: str, file_path: Path) -> tuple[bytes, str]:
    boundary = f"----mc-gate-{uuid.uuid4().hex}"
    newline = b"\r\n"
    chunks: List[bytes] = []
    for key, value in fields.items():
        chunks.append(f"--{boundary}".encode("utf-8"))
        chunks.append(f'Content-Disposition: form-data; name="{key}"'.encode("utf-8"))
        chunks.append(b"")
        chunks.append(str(value).encode("utf-8"))

    file_name = file_path.name
    file_payload = file_path.read_bytes()
    chunks.append(f"--{boundary}".encode("utf-8"))
    chunks.append(
        f'Content-Disposition: form-data; name="{file_field}"; filename="{file_name}"'.encode("utf-8")
    )
    chunks.append(b"Content-Type: text/csv")
    chunks.append(b"")
    chunks.append(file_payload)
    chunks.append(f"--{boundary}--".encode("utf-8"))
    chunks.append(b"")
    body = newline.join(chunks)
    return body, f"multipart/form-data; boundary={boundary}"


def _import_run(base_url: str, csv_file: Path, semester: str, period: str, created_by: str, timeout_sec: int) -> dict:
    body, content_type = _build_multipart(
        fields={"semester": semester, "period": period, "created_by": created_by},
        file_field="file",
        file_path=csv_file,
    )
    payload = _http_json(
        method="POST",
        url=f"{base_url}/api/v1/mc/import/ss01",
        timeout_sec=timeout_sec,
        payload=body,
        headers={"Content-Type": content_type},
    )
    result = payload.get("result") or {}
    run_id = result.get("run_id")
    if not run_id:
        raise RuntimeError(f"Import did not return run_id for period {period}. payload={payload}")
    return payload


def _run_pipeline(base_url: str, run_id: str, created_by: str, timeout_sec: int) -> dict:
    body = json.dumps({"run_id": run_id, "created_by": created_by}, ensure_ascii=False).encode("utf-8")
    return _http_json(
        method="POST",
        url=f"{base_url}/api/v1/mc/run",
        timeout_sec=timeout_sec,
        payload=body,
        headers={"Content-Type": "application/json"},
    )


def _run_checks(base_url: str, run_id: str, created_by: str, timeout_sec: int) -> dict:
    body = json.dumps({"run_id": run_id, "created_by": created_by}, ensure_ascii=False).encode("utf-8")
    return _http_json(
        method="POST",
        url=f"{base_url}/api/v1/mc/checks/run",
        timeout_sec=timeout_sec,
        payload=body,
        headers={"Content-Type": "application/json"},
    )


def _publish_run(base_url: str, run_id: str, created_by: str, timeout_sec: int) -> dict:
    query = url_parse.urlencode({"created_by": created_by})
    return _http_json(
        method="POST",
        url=f"{base_url}/api/v1/mc/runs/{run_id}/publish?{query}",
        timeout_sec=timeout_sec,
    )


def _check_export(base_url: str, run_id: str, created_by: str, export_kind: str, timeout_sec: int) -> dict:
    endpoint = "export.xlsx" if export_kind == "xlsx" else "export.pdf"
    query = url_parse.urlencode({"created_by": created_by})
    data, headers = _http_binary(
        method="GET",
        url=f"{base_url}/api/v1/mc/runs/{run_id}/{endpoint}?{query}",
        timeout_sec=timeout_sec,
    )
    head_hex = data[:8].hex()
    if export_kind == "xlsx":
        signature_ok = head_hex.startswith("504b")
    else:
        signature_ok = head_hex.startswith("25504446")
    return {
        "endpoint": endpoint,
        "size": len(data),
        "content_type": headers.get("content-type"),
        "content_disposition": headers.get("content-disposition"),
        "file_name": _extract_filename_from_disposition(headers.get("content-disposition")),
        "signature_ok": bool(signature_ok),
        "head_hex": head_hex,
    }


def _extract_filename_from_disposition(content_disposition: Optional[str]) -> Optional[str]:
    value = str(content_disposition or "").strip()
    if not value:
        return None
    for part in value.split(";"):
        token = part.strip()
        lower = token.lower()
        if lower.startswith("filename*="):
            raw = token.split("=", 1)[1].strip().strip('"')
            if raw.lower().startswith("utf-8''"):
                raw = raw[7:]
            return url_parse.unquote(raw) or None
        if lower.startswith("filename="):
            return token.split("=", 1)[1].strip().strip('"') or None
    return None


def _list_run_artifacts(
    base_url: str,
    run_id: str,
    timeout_sec: int,
    artifact_type: Optional[str] = None,
) -> List[dict]:
    page = 1
    size = 200
    items: List[dict] = []
    while True:
        query_args = {"page": str(page), "size": str(size)}
        if artifact_type:
            query_args["artifact_type"] = artifact_type
        query = url_parse.urlencode(query_args)
        payload = _http_json(
            method="GET",
            url=f"{base_url}/api/v1/mc/runs/{run_id}/artifacts?{query}",
            timeout_sec=timeout_sec,
        )
        rows = payload.get("items") or []
        items.extend(rows)
        if not payload.get("has_next"):
            break
        page += 1
        if page > 50:
            break
    return items


def _validate_publish_snapshot_totals(
    base_url: str,
    run_id: str,
    timeout_sec: int,
    publish_result: dict,
) -> dict:
    endpoint_by_metric = {
        "halls_rows": "halls",
        "crns_rows": "crns",
        "trainers_rows": "trainers",
        "distribution_rows": "distribution",
    }
    snapshot_totals: Dict[str, int] = {}
    gate_failures: List[str] = []

    for metric_key, endpoint in endpoint_by_metric.items():
        payload = _http_json(
            method="GET",
            url=f"{base_url}/api/v1/mc/runs/{run_id}/{endpoint}?page=1&size=1",
            timeout_sec=timeout_sec,
        )
        total = int(payload.get("total") or 0)
        expected = int(publish_result.get(metric_key) or 0)
        snapshot_totals[metric_key] = total
        if total != expected:
            gate_failures.append(
                f"{metric_key}_snapshot_total_mismatch:{total}!={expected}"
            )

    return {
        "snapshot_totals": snapshot_totals,
        "gate_failures": gate_failures,
    }


def _run_publish_export_regression(
    base_url: str,
    run_id: str,
    created_by: str,
    timeout_sec: int,
) -> dict:
    gate_failures: List[str] = []

    publish_first = _publish_run(
        base_url=base_url,
        run_id=run_id,
        created_by=created_by,
        timeout_sec=timeout_sec,
    )
    publish_second = _publish_run(
        base_url=base_url,
        run_id=run_id,
        created_by=created_by,
        timeout_sec=timeout_sec,
    )

    publish_result_first = publish_first.get("result") or {}
    publish_result_second = publish_second.get("result") or {}
    if str(publish_result_first.get("status")) != "PUBLISHED":
        gate_failures.append(f"publish_status_first={publish_result_first.get('status')}")
    if str(publish_result_second.get("status")) != "PUBLISHED":
        gate_failures.append(f"publish_status_second={publish_result_second.get('status')}")

    snapshot_validation = _validate_publish_snapshot_totals(
        base_url=base_url,
        run_id=run_id,
        timeout_sec=timeout_sec,
        publish_result=publish_result_second or publish_result_first,
    )
    gate_failures.extend(snapshot_validation["gate_failures"])

    xlsx_first = _check_export(
        base_url=base_url,
        run_id=run_id,
        created_by=created_by,
        export_kind="xlsx",
        timeout_sec=timeout_sec,
    )
    xlsx_second = _check_export(
        base_url=base_url,
        run_id=run_id,
        created_by=created_by,
        export_kind="xlsx",
        timeout_sec=timeout_sec,
    )
    pdf_first = _check_export(
        base_url=base_url,
        run_id=run_id,
        created_by=created_by,
        export_kind="pdf",
        timeout_sec=timeout_sec,
    )
    pdf_second = _check_export(
        base_url=base_url,
        run_id=run_id,
        created_by=created_by,
        export_kind="pdf",
        timeout_sec=timeout_sec,
    )

    export_checks = [
        ("xlsx_first", xlsx_first, "spreadsheetml"),
        ("xlsx_second", xlsx_second, "spreadsheetml"),
        ("pdf_first", pdf_first, "application/pdf"),
        ("pdf_second", pdf_second, "application/pdf"),
    ]
    for label, payload, expected_ct in export_checks:
        if not bool(payload.get("signature_ok")):
            gate_failures.append(f"{label}_signature_failed")
        if int(payload.get("size") or 0) <= 0:
            gate_failures.append(f"{label}_empty_payload")
        content_type = str(payload.get("content_type") or "").lower()
        if expected_ct not in content_type:
            gate_failures.append(f"{label}_content_type_invalid={content_type or 'missing'}")
        if not str(payload.get("file_name") or "").strip():
            gate_failures.append(f"{label}_missing_filename")

    artifacts_xlsx = _list_run_artifacts(
        base_url=base_url,
        run_id=run_id,
        timeout_sec=timeout_sec,
        artifact_type="EXPORT_XLSX",
    )
    artifacts_pdf = _list_run_artifacts(
        base_url=base_url,
        run_id=run_id,
        timeout_sec=timeout_sec,
        artifact_type="EXPORT_PDF",
    )
    if len(artifacts_xlsx) < 1:
        gate_failures.append("artifacts_xlsx_missing")
    if len(artifacts_pdf) < 1:
        gate_failures.append("artifacts_pdf_missing")

    xlsx_file_names = {str(item.get("file_name") or "").strip() for item in artifacts_xlsx}
    pdf_file_names = {str(item.get("file_name") or "").strip() for item in artifacts_pdf}
    for expected_name in (
        str(xlsx_first.get("file_name") or "").strip(),
        str(xlsx_second.get("file_name") or "").strip(),
    ):
        if expected_name and expected_name not in xlsx_file_names:
            gate_failures.append(f"xlsx_artifact_not_registered={expected_name}")
    for expected_name in (
        str(pdf_first.get("file_name") or "").strip(),
        str(pdf_second.get("file_name") or "").strip(),
    ):
        if expected_name and expected_name not in pdf_file_names:
            gate_failures.append(f"pdf_artifact_not_registered={expected_name}")

    return {
        "publish_first": publish_first,
        "publish_second": publish_second,
        "export_xlsx_first": xlsx_first,
        "export_xlsx_second": xlsx_second,
        "export_pdf_first": pdf_first,
        "export_pdf_second": pdf_second,
        "snapshot_totals": snapshot_validation["snapshot_totals"],
        "artifact_registry": {
            "xlsx_count": len(artifacts_xlsx),
            "pdf_count": len(artifacts_pdf),
        },
        "gate_failures": gate_failures,
    }


def _run_period_gate(options: GateOptions, csv_file: Path, semester: str, period: str) -> dict:
    period_report: dict[str, Any] = {
        "period": period,
        "status": "PENDING",
        "checks": {},
        "run_id": None,
    }
    created_by = f"{options.created_by}-{period}"

    import_payload = _import_run(
        base_url=options.base_url,
        csv_file=csv_file,
        semester=semester,
        period=period,
        created_by=created_by,
        timeout_sec=options.timeout_sec,
    )
    run_id = str((import_payload.get("result") or {}).get("run_id"))
    period_report["run_id"] = run_id
    period_report["checks"]["import"] = import_payload

    pipeline_payload = _run_pipeline(
        base_url=options.base_url,
        run_id=run_id,
        created_by=created_by,
        timeout_sec=options.timeout_sec,
    )
    period_report["checks"]["pipeline"] = pipeline_payload

    checks_payload = _run_checks(
        base_url=options.base_url,
        run_id=run_id,
        created_by=created_by,
        timeout_sec=options.timeout_sec,
    )
    period_report["checks"]["checks_run"] = checks_payload

    expected = _build_expected_cell_counts(csv_file=csv_file, period=period)
    actual = _build_actual_cell_counts(
        base_url=options.base_url,
        run_id=run_id,
        period=period,
        timeout_sec=options.timeout_sec,
    )
    parity = _compare_count_maps(expected["counts"], actual["counts"])
    period_report["checks"]["distribution_parity"] = {
        "expected_total_loads": expected["total_loads"],
        "actual_total_loads": actual["total_loads"],
        "source_rows_used": expected["source_rows_used"],
        "codes_rows": actual["codes_rows"],
        "rows_used_actual": actual["rows_used"],
        "mismatch_count": parity["mismatch_count"],
        "mismatches_sample": parity["mismatches_sample"],
    }

    excel_parity = _compare_with_evening_excel_cache(
        workbook_file=options.workbook_file,
        counts=actual["counts"],
        period=period,
    )
    period_report["checks"]["excel_cache_parity"] = excel_parity

    publish_export_regression = _run_publish_export_regression(
        base_url=options.base_url,
        run_id=run_id,
        created_by=created_by,
        timeout_sec=options.timeout_sec,
    )
    period_report["checks"]["publish"] = publish_export_regression["publish_first"]
    period_report["checks"]["publish_idempotency"] = publish_export_regression["publish_second"]
    period_report["checks"]["export_xlsx"] = publish_export_regression["export_xlsx_first"]
    period_report["checks"]["export_xlsx_idempotency"] = publish_export_regression["export_xlsx_second"]
    period_report["checks"]["export_pdf"] = publish_export_regression["export_pdf_first"]
    period_report["checks"]["export_pdf_idempotency"] = publish_export_regression["export_pdf_second"]
    period_report["checks"]["publish_export_regression"] = {
        "snapshot_totals": publish_export_regression["snapshot_totals"],
        "artifact_registry": publish_export_regression["artifact_registry"],
        "gate_failures": publish_export_regression["gate_failures"],
    }

    gate_failures = []
    if parity["mismatch_count"] != 0:
        gate_failures.append(f"distribution_parity_mismatch={parity['mismatch_count']}")
    if excel_parity["checked"] and excel_parity["mismatch_count"] != 0:
        gate_failures.append(f"excel_cache_parity_mismatch={excel_parity['mismatch_count']}")

    gate_failures.extend(publish_export_regression["gate_failures"])

    period_report["gate_failures"] = gate_failures
    period_report["status"] = "PASSED" if not gate_failures else "FAILED"
    return period_report


def main() -> int:
    options = parse_args()
    report: dict[str, Any] = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "base_url": options.base_url,
        "workbook_file": str(options.workbook_file),
        "output_file": str(options.output_file),
        "overall_status": "PENDING",
        "period_reports": [],
        "source_csv": None,
    }

    try:
        if options.source_csv:
            source_csv = options.source_csv
            extraction_meta = {"rows_written": None, "semester": None}
        else:
            temp_dir = PROJECT_ROOT / "artifacts" / "acceptance" / "tmp"
            temp_dir.mkdir(parents=True, exist_ok=True)
            source_csv = temp_dir / f"ss01_from_workbook_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            extraction_meta = _extract_ss01_report_to_csv(options.workbook_file, source_csv)
        semester = options.semester or extraction_meta.get("semester")
        if not semester:
            raise RuntimeError("semester could not be resolved from args or workbook SS01_Report!A2")

        report["source_csv"] = str(source_csv)
        report["source_extraction"] = extraction_meta
        report["semester"] = semester

        periods = PERIODS if options.period == "all" else (options.period,)
        for period in periods:
            period_report = _run_period_gate(options, source_csv, semester, period)
            report["period_reports"].append(period_report)

        has_failure = any(item.get("status") != "PASSED" for item in report["period_reports"])
        report["overall_status"] = "FAILED" if has_failure else "PASSED"
    except Exception as exc:  # pragma: no cover - operational path
        report["overall_status"] = "FAILED"
        report["fatal_error"] = str(exc)

    if options.save_timestamped:
        ts_name = f"acceptance_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"
        ts_path = options.output_file.parent / ts_name
        report["timestamped_report"] = str(ts_path)
    options.output_file.parent.mkdir(parents=True, exist_ok=True)
    options.output_file.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    if options.save_timestamped:
        ts_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["overall_status"] == "PASSED" else 1


if __name__ == "__main__":
    raise SystemExit(main())
